import cv2
import pytesseract
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Specify the path to the Tesseract executable
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# Load the pre-trained Haar Cascade for detecting vehicle number plates
plateCascade = cv2.CascadeClassifier("haarcascade_russian_plate_number.xml")
minArea = 500

frameWidth = 640
frameHeight = 480

# Initialize the webcam
cap = cv2.VideoCapture(0)
cap.set(3, frameWidth)
cap.set(4, frameHeight)
cap.set(10, 150)
count = 0

# Define the formats for number plates
formats = [
    r"^[A-Z]{2}\d{2}[A-Z]{2}\d{4}$",  # Format: 2 letters, 2 numbers, 2 letters, 4 numbers
    r"^[A-Z]{3}\d{4}$",               # Format: 3 letters, 4 numbers
    r"^[A-Z]{2}\d{2}[A-Z]{1}\d{4}$",  # Format: 2 letters, 2 numbers, 1 letter, 4 numbers
    r"^\d{2}[A-Z]{2}\d{4}[A-Z]{1}$",  # Format: 2 numbers, 2 letters, 4 numbers, 1 letter
]

# Function to format the number plate
def format_plate(text):
    text = re.sub(r"[^A-Za-z0-9]", "", text)
    for pattern in formats:
        if re.match(pattern, text):
            return text
    return None

# Function to process each frame
def process_frame(img, sheet):
    imgGray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    numberPlates = plateCascade.detectMultiScale(imgGray, 1.1, 4)

    for x, y, w, h in numberPlates:
        area = w * h
        if area > minArea:
            cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
            cv2.putText(img,"NumberPlate",(x, y - 5),cv2.FONT_HERSHEY_COMPLEX,1,(0, 0, 255),2)
            imgRoi = img[y : y + h, x : x + w]

            text = pytesseract.image_to_string(imgRoi, config="--psm 8")
            formatted_text = format_plate(text)
            if formatted_text:
                current_time = datetime.now()
                date_str = current_time.strftime("%Y-%m-%d")
                time_str = current_time.strftime("%H:%M:%S")
                print(f"Formatted Text: {formatted_text} at {date_str} {time_str}")
                sheet.append([formatted_text, date_str, time_str])

    return img

# Main function
def main():
    file_path = "NumberPlates.xlsx"

    # Load the existing workbook or create a new one if it doesn't exist
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Number Plates"
        sheet.append(["Number Plate", "Date", "Time"])

    while True:
        success, img = cap.read()
        if not success:
            break

        img = process_frame(img, sheet)
        cv2.imshow("Result", img)

        # Allow user to press 's' to save the image
        if cv2.waitKey(1) & 0xFF == ord("s"):
            cv2.imwrite(f".\\IMAGES\\{str(count)}.jpg", img)
            cv2.rectangle(img, (0, 200), (640, 300), (0, 255, 0), cv2.FILLED)
            cv2.putText(img,"Scan Saved",(15, 265),cv2.FONT_HERSHEY_COMPLEX,2,(0, 0, 255),2,)
            cv2.imshow("Result", img)
            cv2.waitKey(500)
            count += 1

        if cv2.getWindowProperty("Result", cv2.WND_PROP_VISIBLE) < 1:
            break

    cap.release()
    cv2.destroyAllWindows()
    workbook.save(file_path)

if __name__ == "__main__":
    main()